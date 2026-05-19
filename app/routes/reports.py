"""
Routes: Relatórios
"""
import io
from datetime import date
from flask import Blueprint, render_template, request, send_file, flash
from flask_login import login_required, current_user
from sqlalchemy import extract, func, or_
from app import db
from app.models.expense  import Expense
from app.models.income   import Income
from app.models.category import Category
from app.services.finance_service import FinanceService

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/')
@login_required
def index():
    today = date.today()
    year  = int(request.args.get('year',  today.year))
    month = int(request.args.get('month', today.month))

    summary  = FinanceService.get_monthly_summary(current_user.id, year, month)
    cats     = FinanceService.get_expenses_by_category(current_user.id, year, month)
    chart    = FinanceService.get_monthly_chart_data(current_user.id, months=12)
    health   = FinanceService.get_financial_health(current_user.id)
    forecast = FinanceService.get_expense_forecast(current_user.id)

    return render_template(
        'reports/index.html',
        summary  = summary,
        cats     = cats,
        chart    = chart,
        health   = health,
        forecast = forecast,
        year     = year,
        month    = month,
    )


@reports_bp.route('/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    year  = int(request.args.get('year',  today.year))
    month = int(request.args.get('month', today.month))

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Despesas'
    ws2 = wb.create_sheet('Receitas')
    ws3 = wb.create_sheet('Resumo')

    # ── Estilo ────────────────────────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    month_names = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                   'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border    = thin_border

    # ── Despesas ──────────────────────────────────────────────────────────────
    style_header(ws1, ['Data', 'Nome', 'Categoria', 'Valor (R$)', 'Forma Pagamento', 'Observação'])
    expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        extract('year',  Expense.date) == year,
        extract('month', Expense.date) == month,
    ).order_by(Expense.date).all()

    total_exp = 0
    for e in expenses:
        cat_name = e.category.name if e.category else 'Sem categoria'
        ws1.append([
            e.date.strftime('%d/%m/%Y'), e.name, cat_name,
            float(e.amount), e.payment_method, e.notes or ''
        ])
        total_exp += float(e.amount)
    ws1.append(['', '', 'TOTAL', total_exp, '', ''])

    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ── Receitas ──────────────────────────────────────────────────────────────
    style_header(ws2, ['Data', 'Nome', 'Tipo', 'Valor (R$)', 'Observação'])
    incomes = Income.query.filter(
        Income.user_id == current_user.id,
        extract('year',  Income.date) == year,
        extract('month', Income.date) == month,
    ).order_by(Income.date).all()

    total_inc = 0
    for i in incomes:
        ws2.append([i.date.strftime('%d/%m/%Y'), i.name, i.type, float(i.amount), i.notes or ''])
        total_inc += float(i.amount)
    ws2.append(['', 'TOTAL', '', total_inc, ''])

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ── Resumo ────────────────────────────────────────────────────────────────
    ws3.append(['Finance Control Pro - Relatório Financeiro'])
    ws3['A1'].font = Font(size=14, bold=True, color='4F46E5')
    ws3.append([f'Período: {month_names[month-1]}/{year}'])
    ws3.append([])
    ws3.append(['Indicador', 'Valor'])
    style_header(ws3, ['Indicador', 'Valor'])

    # override row 4
    ws3['A4'] = 'Indicador'; ws3['B4'] = 'Valor'
    ws3['A4'].font = header_font; ws3['A4'].fill = header_fill
    ws3['B4'].font = header_font; ws3['B4'].fill = header_fill

    ws3.append(['Total Receitas',  f'R$ {total_inc:,.2f}'])
    ws3.append(['Total Despesas',  f'R$ {total_exp:,.2f}'])
    ws3.append(['Saldo',           f'R$ {total_inc - total_exp:,.2f}'])
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'finance_{year}_{month:02d}.xlsx'
    return send_file(output, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)


@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    today = date.today()
    year  = int(request.args.get('year',  today.year))
    month = int(request.args.get('month', today.month))
    month_names = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                   'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 textColor=colors.HexColor('#4F46E5'), fontSize=20)
    elements.append(Paragraph('Finance Control Pro', title_style))
    elements.append(Paragraph(f'Relatório - {month_names[month-1]}/{year}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    # Resumo
    summary = FinanceService.get_monthly_summary(current_user.id, year, month)
    sum_data = [
        ['Indicador', 'Valor'],
        ['Total Receitas',  f"R$ {summary['total_incomes']:,.2f}"],
        ['Total Despesas',  f"R$ {summary['total_expenses']:,.2f}"],
        ['Saldo',           f"R$ {summary['balance']:,.2f}"],
    ]
    t = Table(sum_data, colWidths=[8*cm, 8*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f0f0')]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN',      (1,0), (1,-1), 'RIGHT'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))

    # Despesas
    elements.append(Paragraph('Despesas do Período', styles['Heading2']))
    expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        extract('year',  Expense.date) == year,
        extract('month', Expense.date) == month,
    ).order_by(Expense.date).all()

    exp_data = [['Data', 'Nome', 'Categoria', 'Valor']]
    for e in expenses:
        exp_data.append([
            e.date.strftime('%d/%m/%Y'), e.name[:30],
            e.category.name if e.category else '-',
            f"R$ {float(e.amount):,.2f}"
        ])

    if len(exp_data) > 1:
        te = Table(exp_data, colWidths=[3*cm, 7*cm, 4*cm, 3*cm])
        te.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#ffeaea')]),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.grey),
        ]))
        elements.append(te)

    doc.build(elements)
    buffer.seek(0)

    filename = f'relatorio_{year}_{month:02d}.pdf'
    return send_file(buffer, download_name=filename,
                     mimetype='application/pdf', as_attachment=True)